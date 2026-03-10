# -*- coding: utf-8 -*-
"""
core/excel_utils.py — Utilidades Excel centralizadas

Regla de oro:
  - LECTURA  → pandas read_excel  (rápido, no modifica el archivo)
  - ESCRITURA de archivos nuevos / reemplazados totalmente
              → pd.ExcelWriter con engine="openpyxl"  (crea archivo limpio)
  - ESCRITURA de archivos pre-existentes donde solo hay que tocar
    algunas celdas (origenDestino, plantillas Bancolombia)
              → load_workbook + wb.save()  (preserva formatos y fechas)

Funciones públicas
──────────────────
read_sheet_safe(path, sheet_name, context)  → pd.DataFrame
write_sheets(path, sheets, context)         → None
write_cells(path, sheet_name, cell_updates) → None   ← nuevo
ensure_columns(df, cols)                    → pd.DataFrame
reorder_columns(df, cols)                   → pd.DataFrame
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd

# openpyxl se importa localmente donde se necesita para evitar
# cargarlo si solo se usa read_sheet_safe

__all__ = [
    "read_sheet_safe",
    "write_sheets",
    "write_cells",
    "ensure_columns",
    "reorder_columns",
]


# =========================================================
# LECTURA
# =========================================================
def read_sheet_safe(
    path: Path,
    sheet_name: str,
    context: str = "",
    dtype: Optional[Dict] = None,
) -> pd.DataFrame:
    """
    Lee una hoja de Excel de forma segura.

    - Si el archivo no existe → retorna DataFrame vacío (no lanza error).
    - Si la hoja no existe   → retorna DataFrame vacío (no lanza error).
    - Siempre normaliza nombres de columnas (quita espacios y nbsp).

    Usa pd.read_excel (solo lectura → no toca el archivo, sin riesgo de
    alterar formatos).
    """
    tag = f"[{context}] " if context else ""

    if not Path(path).exists():
        return pd.DataFrame()

    try:
        df = pd.read_excel(
            path,
            sheet_name=sheet_name,
            dtype=dtype,
            engine=_engine_for(path),
        )
    except Exception as e:
        err = str(e).lower()
        if "worksheet" in err or "sheet" in err or "no sheet" in err:
            return pd.DataFrame()
        raise RuntimeError(
            f"{tag}No se pudo leer '{sheet_name}' de {Path(path).name}: {e}"
        ) from e

    # Normalizar encabezados
    df.columns = [str(c).replace("\u00A0", " ").strip() for c in df.columns]
    return df


def _engine_for(path: Path) -> Optional[str]:
    """Devuelve el engine correcto según la extensión."""
    suffix = Path(path).suffix.lower()
    if suffix == ".xlsb":
        return "pyxlsb"
    if suffix in (".xlsx", ".xlsm", ".xltx"):
        return "openpyxl"
    return None  # pandas elige automáticamente


# =========================================================
# ESCRITURA COMPLETA (archivos nuevos o reemplazados totalmente)
# =========================================================
def write_sheets(
    path: Path,
    sheets: Dict[str, pd.DataFrame],
    context: str = "",
) -> None:
    """
    Escribe uno o varios DataFrames en un archivo Excel.

    Comportamiento:
      - Si el archivo no existe → lo crea.
      - Si el archivo existe    → reemplaza SOLO las hojas indicadas,
        preserva las demás hojas.
      - Engine siempre openpyxl (no xlwt, no xlsxwriter).
      - Columnas normalizadas en los encabezados.

    Uso:
        write_sheets(
            config.SWIFT_COMPLETOS,
            {config.SHEET_V1: df_v1, config.SHEET_V2: df_v2},
            context="completos",
        )

    ⚠ Esta función reescribe las hojas indicadas completamente.
      Para modificar celdas individuales en un archivo pre-existente
      (preservando formatos de fecha, números, etc.) usar write_cells().
    """
    tag  = f"[{context}] " if context else ""
    path = Path(path)

    path.parent.mkdir(parents=True, exist_ok=True)

    try:
        mode           = "a" if path.exists() else "w"
        if_sheet_exists = "replace" if mode == "a" else None

        writer_kwargs = dict(engine="openpyxl", mode=mode)
        if if_sheet_exists:
            writer_kwargs["if_sheet_exists"] = if_sheet_exists

        with pd.ExcelWriter(path, **writer_kwargs) as writer:
            for sheet_name, df in sheets.items():
                safe_df = df.copy()
                safe_df.columns = [
                    str(c).replace("\u00A0", " ").strip()
                    for c in safe_df.columns
                ]
                safe_df.to_excel(writer, sheet_name=sheet_name, index=False)

    except PermissionError:
        raise PermissionError(
            f"\n{'='*60}\n"
            f"  {tag}No se pudo guardar {path.name} — está abierto en Excel.\n"
            f"  Cerralo y volvé a ejecutar.\n"
            f"  Ruta: {path}\n"
            f"{'='*60}"
        )


# =========================================================
# ESCRITURA DE CELDAS INDIVIDUALES (archivos pre-existentes)
# =========================================================
def write_cells(
    path: Path,
    sheet_name: str,
    cell_updates: List[Tuple[int, int, object]],
    create_header_if_missing: Optional[Tuple[int, int, str]] = None,
    context: str = "",
) -> None:
    """
    Escribe SOLO las celdas indicadas en un archivo Excel pre-existente.

    Usa load_workbook + wb.save() → preserva 100% formatos, fechas
    y tipos de datos del resto del archivo.

    Parámetros:
        path            : ruta al archivo Excel
        sheet_name      : nombre de la hoja a modificar
        cell_updates    : lista de (row, col, value)  — 1-indexado
        create_header_if_missing:
                          (row, col, header_text) — si la columna no existe,
                          la crea con ese encabezado en esa celda.
        context         : string para mensajes de error/log

    Ejemplo:
        write_cells(
            config.ORIGEN_DESTINO,
            "Origen y destino",
            [(15, 11, "mi-llave-abc123"), (16, 11, "otra-llave")],
        )

    ⚠ No usar para archivos que se crean de cero; usar write_sheets() en su lugar.
    """
    from openpyxl import load_workbook

    tag  = f"[{context}] " if context else ""
    path = Path(path)

    if not path.exists():
        raise FileNotFoundError(f"{tag}Archivo no encontrado: {path}")

    try:
        wb = load_workbook(path)
    except PermissionError:
        raise PermissionError(
            f"\n{'='*60}\n"
            f"  {tag}No se pudo abrir {path.name} — está abierto en Excel.\n"
            f"  Cerralo y volvé a ejecutar.\n"
            f"  Ruta: {path}\n"
            f"{'='*60}"
        )

    if sheet_name not in wb.sheetnames:
        raise KeyError(
            f"{tag}Hoja '{sheet_name}' no encontrada en {path.name}. "
            f"Hojas disponibles: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    if create_header_if_missing:
        h_row, h_col, h_text = create_header_if_missing
        if ws.cell(row=h_row, column=h_col).value is None:
            ws.cell(row=h_row, column=h_col, value=h_text)

    for row, col, value in cell_updates:
        ws.cell(row=row, column=col, value=value)

    try:
        wb.save(path)
    except PermissionError:
        raise PermissionError(
            f"\n{'='*60}\n"
            f"  {tag}No se pudo guardar {path.name} — está abierto en Excel.\n"
            f"  Cerralo y volvé a ejecutar.\n"
            f"  Ruta: {path}\n"
            f"{'='*60}"
        )


# =========================================================
# UTILIDADES DE DataFrame
# =========================================================
def ensure_columns(df: pd.DataFrame, cols: Sequence[str]) -> pd.DataFrame:
    """
    Garantiza que el DataFrame tenga todas las columnas en `cols`.
    Las que falten se agregan vacías (pd.NA).
    Las columnas extras se conservan.
    """
    out = df.copy()
    for col in cols:
        if col not in out.columns:
            out[col] = pd.NA
    return out


def reorder_columns(df: pd.DataFrame, cols: Sequence[str]) -> pd.DataFrame:
    """
    Reordena el DataFrame para que las columnas sigan el orden de `cols`.

    - Las columnas de `cols` que faltan en df se agregan vacías (pd.NA).
    - Las columnas de df que no están en `cols` se agregan al final.
    """
    out = df.copy()
    for col in cols:
        if col not in out.columns:
            out[col] = pd.NA

    extras = [c for c in out.columns if c not in cols]
    return out[list(cols) + extras]


# =========================================================
# LEER COLUMNAS DE UN XLSX SIN CARGAR TODOS LOS DATOS
# =========================================================
def get_sheet_headers(path: Path, sheet_name: str) -> Dict[str, int]:
    """
    Retorna {nombre_columna: numero_columna_1indexed} leyendo solo la fila 1.
    Usa openpyxl directamente → más rápido que leer todo el sheet con pandas.
    Normaliza headers (quita nbsp y espacios).
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise KeyError(f"Hoja '{sheet_name}' no encontrada en {Path(path).name}.")

    ws   = wb[sheet_name]
    hdrs = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None:
            key = str(cell.value).replace("\u00A0", " ").strip()
            hdrs[key] = cell.column
    wb.close()
    return hdrs