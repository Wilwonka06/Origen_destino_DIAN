# -*- coding: utf-8 -*-
"""
run_formulario.py — Cruces de Formulario, Llave y Llave OD

PASO 1) Lee XLSB (hoja COM)
        - Encabezados desde fila 4
        - Toma A..F (hasta DEBITO)

PASO 2) Filtra:
        - FECHA >= config.FECHA_MIN_XLSB
        - INDICA == Imp

PASO 3) Cruza con Swift_completos:
        - Llave 1: FECHA_COM (normalizada a YYYY-MM-DD) vs Date (Swift)
        - Llave 2: DETALLE_COM (limpia desde #) vs Nombre archivo (Swift)
                 (Nombre archivo: elimina códigos iniciales + elimina .pdf)
                 Matching robusto por tokens:
                   * exige coincidencia de primeras 2 palabras (si existen)
                   * exige ratio de coincidencia >= TOKEN_MIN_RATIO
        - Trae FORMULARIO -> Formulario (Swift)
        - Si varios matches, suma DEBITO y si coincide con Amount, concatena formularios con "-"

PASO 4) Cruce Llave (origenDestino.xlsx):
        - Llave: "Nombre personalizado" (origenDestino / hoja Datos Origen Destino)
                 vs "Nombre personalizado" (Swift)
        - Trae: "Llave carga masiva" -> "Llave" (Swift)

PASO 5) Cruce final (origenDestino.xlsx / hoja "Origen y destino"):
        - Desde Swift: columna "Formulario" (ej: "12030-None-12028")
          * extrae consecutivos numéricos
        - Relaciona con origenDestino: columna "Consecutivo"
        - Escribe Swift["Llave"] en origenDestino["Llave Origen Destino"]
        - Guarda ambos archivos
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List

import pandas as pd
from pyxlsb import open_workbook

import config
from core.logger import get_logger
from core.validators import validate_input_files
from core.excel_utils import write_sheets, read_sheet_safe

LOGGER = get_logger(__name__)

# ─── constantes ────────────────────────────────────────────
FECHA_MIN         = pd.Timestamp(config.FECHA_MIN_XLSB)
AMOUNT_TOL        = config.AMOUNT_TOL
TOKEN_MIN_RATIO   = config.TOKEN_MIN_RATIO
TOKEN_MIN_OVERLAP = config.TOKEN_MIN_OVERLAP
HEADER_ROW_1BASED = 4
MAX_COLS          = 6


# =========================================================
# UTILIDADES NORMALIZACIÓN
# =========================================================
def _parse_fecha_excel_series(s: pd.Series) -> pd.Series:
    if s is None:
        return pd.to_datetime(pd.Series([], dtype="datetime64[ns]"))
    if pd.api.types.is_datetime64_any_dtype(s):
        return pd.to_datetime(s, errors="coerce")
    if pd.api.types.is_numeric_dtype(s):
        return pd.to_datetime(s, unit="D", origin="1899-12-30", errors="coerce")
    s2 = s.astype(str).str.strip()
    s2 = s2.replace({"": pd.NA, "nan": pd.NA, "NaT": pd.NA})
    return pd.to_datetime(s2, dayfirst=True, errors="coerce")


def _normalize_text_key(x) -> str:
    """Normaliza texto: lowercase, sin espacios extra, sin nbsp."""
    if x is None:
        return ""
    s = str(x).replace("\u00A0", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s.casefold()


def _tokenize(s: str) -> List[str]:
    if not s:
        return []
    return re.findall(r"[a-z0-9]+", _normalize_text_key(s))


def _clean_detalle(detalle) -> str:
    """Extrae el texto ANTES del '#' en DETALLE de COM y normaliza."""
    if detalle is None:
        return ""
    s = str(detalle).replace("\u00A0", " ")
    s = re.split(r"#", s, maxsplit=1)[0].strip()
    return _normalize_text_key(s)


def _clean_nombre_archivo(nombre) -> str:
    """
    Limpia el campo 'Nombre archivo' de Swift:
      - Quita extensión .pdf
      - Quita prefijos numéricos iniciales (ej: '11487 11453 ')
      - Normaliza espacios y lowercase
    """
    if nombre is None:
        return ""
    s = str(nombre).replace("\u00A0", " ").strip()
    s = re.sub(r"\.pdf\s*$", "", s, flags=re.IGNORECASE).strip()
    s = re.sub(r"^(?:\d+\s+)+", "", s).strip()
    s = re.sub(r"\s+", " ", s).strip()
    return _normalize_text_key(s)


def _parse_money_to_float(v) -> float:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return float("nan")
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).replace("\u00A0", " ").strip()
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^0-9\.,\-]", "", s)
    if s in ("", "-", ".", ","):
        return float("nan")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    if re.search(r"\.$", s):
        s += "0"
    try:
        return float(s)
    except Exception:
        return float("nan")


# =========================================================
# MATCH ROBUSTO (tokens)
# =========================================================
def _tokens_match(swift_clean: str, detalle_clean: str) -> bool:
    """
    True si swift_clean coincide con detalle_clean por tokens.
    - nombre >= 2 tokens: exige las primeras 2 palabras + ratio >= TOKEN_MIN_RATIO
    - nombre 1 token: ese token debe estar en detalle
    """
    st = _tokenize(swift_clean)
    dt = set(_tokenize(detalle_clean))

    if not st:
        return False

    overlap = sum(1 for t in st if t in dt)
    ratio   = overlap / max(len(st), 1)

    if len(st) >= 2:
        if st[0] not in dt or st[1] not in dt:
            return False
        if overlap < TOKEN_MIN_OVERLAP:
            return False
        if ratio < TOKEN_MIN_RATIO:
            return False
        return True

    return st[0] in dt


# =========================================================
# PASO 1) LECTURA XLSB COM (A..F, header fila 4)
# =========================================================
def read_com_sheet(xlsb_path: Path, sheet_name: str = config.SHEET_COM) -> pd.DataFrame:
    if not xlsb_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {xlsb_path}")

    header_idx = HEADER_ROW_1BASED - 1
    data_rows: List[List] = []
    headers = None

    with open_workbook(str(xlsb_path)) as wb:
        with wb.get_sheet(sheet_name) as sheet:
            for r_idx, row in enumerate(sheet.rows()):
                if r_idx < header_idx:
                    continue
                values = [cell.v for cell in row[:MAX_COLS]]
                if r_idx == header_idx:
                    headers = [str(v).strip() if v is not None else "" for v in values]
                    continue
                if headers is None:
                    raise RuntimeError("No se detectó header en fila 4 del XLSB.")
                if all(v is None or str(v).strip() == "" for v in values):
                    continue
                data_rows.append(values)

    df = pd.DataFrame(data_rows, columns=headers)
    df.columns = [str(c).replace("\u00A0", " ").strip() for c in df.columns]
    LOGGER.info(f"XLSB leído → filas={len(df)} | cols={list(df.columns)}")
    return df


# =========================================================
# PASO 2) FILTROS FECHA + INDICA=Imp
# =========================================================
def filter_com_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        LOGGER.warning("COM viene vacío antes de filtrar.")
        return df

    cols_map = {str(c).strip().upper(): c for c in df.columns}

    if "FECHA" not in cols_map:
        raise KeyError(f"No encuentro columna FECHA. Columnas: {list(df.columns)}")
    if "INDICA" not in cols_map:
        raise KeyError(f"No encuentro columna INDICA. Columnas: {list(df.columns)}")

    out = df.copy()
    out["_FECHA_DT"] = _parse_fecha_excel_series(out[cols_map["FECHA"]])
    out["_INDICA_NORM"] = (
        out[cols_map["INDICA"]]
        .astype(str)
        .str.replace("\u00A0", " ", regex=False)
        .str.strip()
        .str.lower()
    )

    before = len(out)
    out = out.loc[out["_FECHA_DT"].notna()].copy()
    out = out.loc[out["_FECHA_DT"] >= FECHA_MIN].copy()
    out = out.loc[out["_INDICA_NORM"] == "imp"].copy()
    out = out.drop(columns=["_FECHA_DT", "_INDICA_NORM"], errors="ignore").reset_index(drop=True)

    LOGGER.info(
        f"Filtro COM: inicio={before} → resultado={len(out)} "
        f"(FECHA>={FECHA_MIN.date()}, INDICA=Imp)"
    )
    return out


# =========================================================
# PASO 3) CRUCE CON SWIFT_COMPLETOS → FORMULARIO
# =========================================================
def _build_com_keys(df_com: pd.DataFrame) -> pd.DataFrame:
    """
    Prepara COM con columnas para el cruce:
      fecha_key      → YYYY-MM-DD
      detalle_clean  → DETALLE antes del '#', normalizado
      debito_num     → DEBITO como float
      formulario_str → FORMULARIO como string (tal cual, sin lstrip)
      row_order      → orden original
    """
    cols_map = {str(c).strip().upper(): c for c in df_com.columns}
    required = ["FECHA", "DETALLE", "FORMULARIO", "DEBITO"]
    missing  = [c for c in required if c not in cols_map]
    if missing:
        raise KeyError(f"En COM faltan columnas: {missing}. Detectadas: {list(df_com.columns)}")

    out = df_com.copy()
    out["_fecha_dt"]      = _parse_fecha_excel_series(out[cols_map["FECHA"]])
    out["fecha_key"]      = out["_fecha_dt"].dt.strftime("%Y-%m-%d")
    out["detalle_clean"]  = out[cols_map["DETALLE"]].apply(_clean_detalle)
    out["debito_num"]     = out[cols_map["DEBITO"]].apply(_parse_money_to_float)
    out["formulario_str"] = out[cols_map["FORMULARIO"]].apply(
        lambda x: ""
        if x is None or (isinstance(x, float) and pd.isna(x))
        else re.sub(r"\.0+$", "", str(x).strip()).lstrip("0") or "0"
    )
    out["row_order"] = range(len(out))
    out = out.drop(columns=["_fecha_dt"], errors="ignore")
    return out


def _build_swift_keys(df_swift: pd.DataFrame) -> pd.DataFrame:
    """
    Prepara Swift con columnas para el cruce:
      fecha_key    → YYYY-MM-DD
      nombre_clean → Nombre archivo limpio
      amount_num   → Amount como float
    """
    out  = df_swift.copy()
    miss = [c for c in ("Date", "Nombre archivo", "Amount", "id") if c not in out.columns]
    if miss:
        raise KeyError(f"Swift_completos falta(n) columna(s) {miss}. Detectadas: {list(out.columns)}")

    out["_date_dt"]     = pd.to_datetime(out["Date"], errors="coerce")
    out["fecha_key"]    = out["_date_dt"].dt.strftime("%Y-%m-%d")
    out["nombre_clean"] = out["Nombre archivo"].apply(_clean_nombre_archivo)
    out["amount_num"]   = out["Amount"].apply(_parse_money_to_float)
    out = out.drop(columns=["_date_dt"], errors="ignore")
    return out


def _update_formulario_for_sheet(
    df_swift_sheet: pd.DataFrame,
    df_com: pd.DataFrame,
) -> pd.DataFrame:
    """
    Cruza COM → Swift para poblar 'Formulario'.

    Lógica (fiel al archivo de referencia):
      1. Para cada Swift, busca filas COM con la misma fecha
      2. Filtra candidatos por match de tokens (Nombre archivo vs DETALLE)
      3. 1 candidato  → asigna FORMULARIO directamente
      4. N candidatos → si sum(DEBITO) ≈ Amount → concatena formularios con '-'
                        si no cuadra → no asigna
    """
    if df_swift_sheet.empty:
        return df_swift_sheet

    out = df_swift_sheet.copy()
    out.columns = [str(c).strip() for c in out.columns]

    if "Formulario" not in out.columns:
        out["Formulario"] = ""

    com_k   = _build_com_keys(df_com)
    swift_k = _build_swift_keys(out)

    # Índice COM por fecha para búsqueda rápida
    com_by_date: Dict[str, pd.DataFrame] = {
        k: v.copy() for k, v in com_k.groupby("fecha_key")
    }

    updated       = 0
    multi_matched = 0
    multi_ok      = 0
    multi_fail    = 0

    for _, srow in swift_k.iterrows():
        sid      = srow["id"]
        s_fecha  = srow["fecha_key"]
        s_name   = srow["nombre_clean"]
        s_amount = srow["amount_num"]

        if not isinstance(s_fecha, str) or not s_fecha:
            continue

        com_day = com_by_date.get(s_fecha)
        if com_day is None or com_day.empty:
            continue

        # Filtrar por match de tokens
        cand = com_day.loc[
            com_day["detalle_clean"].apply(lambda d: _tokens_match(s_name, d))
        ].copy()

        if cand.empty:
            continue

        if len(cand) == 1:
            form_val = str(cand.iloc[0]["formulario_str"]).strip()
            if form_val and form_val.lower() != "none":
                out.loc[out["id"] == sid, "Formulario"] = form_val
                updated += 1
            continue

        # Múltiples candidatos → verificar suma de débitos
        multi_matched += 1
        deb_sum = cand["debito_num"].sum(skipna=True)

        if pd.notna(deb_sum) and pd.notna(s_amount) and abs(deb_sum - s_amount) <= AMOUNT_TOL:
            cand  = cand.sort_values("row_order")
            forms = [
                str(x).strip()
                for x in cand["formulario_str"].tolist()
                if str(x).strip() not in ("", "none", "nan", "nat")
            ]
            if forms:
                out.loc[out["id"] == sid, "Formulario"] = "-".join(forms)
                updated += 1
                multi_ok += 1
        else:
            multi_fail += 1

    LOGGER.info(
        f"Swift actualizado Formulario: updated={updated} | "
        f"multi_matched={multi_matched} | multi_ok={multi_ok} | multi_fail={multi_fail}"
    )
    return out


# =========================================================
# PASO 4) CRUCE ORIGEN DESTINO → LLAVE (Swift)
# =========================================================
def _read_od_mapping(path: Path) -> pd.DataFrame:
    """Lee Nombre personalizado → Llave carga masiva desde origenDestino."""
    if not path.exists():
        raise FileNotFoundError(f"No existe origenDestino.xlsx: {path}")

    df = pd.read_excel(path, sheet_name=config.SHEET_OD_DATOS)
    df.columns = [str(c).replace("\u00A0", " ").strip() for c in df.columns]

    for col in (config.OD_COL_NOMBRE, config.OD_COL_LLAVE):
        if col not in df.columns:
            raise KeyError(
                f"No se encontró columna '{col}' en {path.name}. "
                f"Columnas disponibles: {list(df.columns)}"
            )

    out = df[[config.OD_COL_NOMBRE, config.OD_COL_LLAVE]].copy()
    out[config.OD_COL_NOMBRE] = out[config.OD_COL_NOMBRE].apply(_normalize_text_key)
    out[config.OD_COL_LLAVE]  = out[config.OD_COL_LLAVE].astype(str).str.strip()

    out = (
        out
        .loc[out[config.OD_COL_NOMBRE] != ""]
        .loc[out[config.OD_COL_LLAVE]  != ""]
        .drop_duplicates(subset=[config.OD_COL_NOMBRE], keep="first")
        .reset_index(drop=True)
    )

    LOGGER.info(f"OrigenDestino mapping cargado: {len(out)} llaves únicas.")
    return out


def _apply_llave_to_sheet(
    df_swift_sheet: pd.DataFrame,
    od_map: pd.DataFrame,
) -> pd.DataFrame:
    """
    Aplica cruce Llave a una hoja Swift.

    El 'Nombre personalizado' en Swift está truncado a 50 caracteres,
    por lo que no coincide exactamente con origenDestino.
    Ejemplo:
      Swift:       'GESI TEKSTIL ITHALAT IHRACAT TICARE PNBPUS3NNYC'
      origenDestino: 'GESI TEKSTIL ITHALAT IHRACAT PNBPUS3NNYC'

    Estrategia: match por tokens con las mismas reglas que Paso 3
      - Las primeras 2 palabras del Swift deben estar en origenDestino
      - Token overlap ratio >= TOKEN_MIN_RATIO
    """
    if df_swift_sheet.empty:
        return df_swift_sheet

    out = df_swift_sheet.copy()
    out.columns = [str(c).strip() for c in out.columns]

    if config.OD_COL_NOMBRE not in out.columns:
        raise KeyError("Swift_completos no tiene columna 'Nombre personalizado'.")
    if "Llave" not in out.columns:
        out["Llave"] = ""

    # Preparar lista de (tokens_od, llave) para búsqueda por tokens
    od_entries = [
        (_tokenize(nombre), llave)
        for nombre, llave in zip(od_map[config.OD_COL_NOMBRE], od_map[config.OD_COL_LLAVE])
        if nombre and llave
    ]

    def _find_llave(nombre_swift_norm: str) -> str:
        """Busca la llave en origenDestino por token overlap."""
        if not nombre_swift_norm:
            return ""

        st = _tokenize(nombre_swift_norm)
        if not st:
            return ""

        best_llave = ""
        best_ratio = 0.0

        for od_tokens, llave in od_entries:
            od_set  = set(od_tokens)
            overlap = sum(1 for t in st if t in od_set)
            ratio   = overlap / max(len(st), len(od_tokens), 1)

            # Exige primeras 2 palabras coincidan (igual que _tokens_match)
            if len(st) >= 2 and (st[0] not in od_set or st[1] not in od_set):
                continue
            if overlap < TOKEN_MIN_OVERLAP:
                continue
            if ratio < TOKEN_MIN_RATIO:
                continue

            if ratio > best_ratio:
                best_ratio = ratio
                best_llave = llave

        return best_llave

    out["_np_norm"]   = out[config.OD_COL_NOMBRE].apply(_normalize_text_key)
    out["_llave_cur"] = out["Llave"].fillna("").astype(str).str.strip()

    before_empty = (out["_llave_cur"] == "").sum()
    filled = 0
    sin_match_nombres = []

    for idx, row in out.iterrows():
        if row["_llave_cur"] != "":
            continue  # ya tiene llave, no sobreescribir

        llave_nueva = _find_llave(row["_np_norm"])

        if llave_nueva:
            out.at[idx, "Llave"] = llave_nueva
            filled += 1
        else:
            sin_match_nombres.append(row["_np_norm"])

    after_empty = (out["Llave"].fillna("").astype(str).str.strip() == "").sum()
    LOGGER.info(
        f"Cruce Llave aplicado: filled={filled} | "
        f"empty_before={before_empty} | empty_after={after_empty}"
    )

    if sin_match_nombres:
        unicos = list(dict.fromkeys(sin_match_nombres))  # dedup preservando orden
        LOGGER.warning(
            f"Sin llave ({len(unicos)} 'Nombre personalizado' únicos sin match):\n"
            + "\n".join(f"  '{n}'" for n in unicos[:20])
        )

    out = out.drop(columns=["_np_norm", "_llave_cur"], errors="ignore")
    return out


# =========================================================
# PASO 5) FORMULARIO (Swift) → CONSECUTIVO (origenDestino)
# =========================================================
def _extract_consecutivos_from_formulario(val) -> List[str]:
    """
    '12030-None-12028' → ['12030', '12028']
    Ignora None / vacíos, extrae solo dígitos.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return []
    s = str(val).strip()
    if not s:
        return []
    parts = [p.strip() for p in s.split("-")]
    out = []
    for p in parts:
        if not p or p.lower() == "none":
            continue
        m = re.search(r"\d+", p)
        if m:
            norm = m.group(0).lstrip("0") or "0"
            out.append(norm)
    return out


def _normalize_consecutivo_series(s: pd.Series) -> pd.Series:
    """Normaliza Consecutivo a string de dígitos sin ceros iniciales (12029.0 → '12029')."""
    def _one(x):
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return ""
        if isinstance(x, int) and not isinstance(x, bool):
            return str(x).lstrip("0") or "0"
        if isinstance(x, float):
            if pd.isna(x):
                return ""
            if float(x).is_integer():
                return str(int(x)).lstrip("0") or "0"
            return str(x).strip()
        sx = str(x).strip()
        sx = re.sub(r"\.0+$", "", sx)
        m  = re.search(r"\d+", sx)
        return (m.group(0).lstrip("0") or "0") if m else ""
    return s.apply(_one)


def _update_od_llave(path: Path, df_swift_all: pd.DataFrame) -> None:
    """
    Actualiza SOLO la columna 'Llave Origen Destino' en origenDestino.

    Usa openpyxl directamente para escribir únicamente las celdas que cambian,
    sin tocar fechas, números ni formato de ninguna otra columna.

    Lógica:
      1. Lee Swift.Formulario, separa por '-' → lista de formularios
      2. Para cada formulario busca en Consecutivo de origenDestino
      3. Si coincide → escribe la Llave en 'Llave Origen Destino' de esa fila
    """
    if not path.exists():
        raise FileNotFoundError(f"No existe origenDestino.xlsx: {path}")

    if df_swift_all.empty:
        LOGGER.info("PASO 5 → Swift vacío. No hay nada que cruzar hacia origenDestino.")
        return

    for col in ("Formulario", "Llave"):
        if col not in df_swift_all.columns:
            raise KeyError(f"Swift_completos no tiene columna '{col}' requerida para PASO 5.")

    # ── Construir mapping formulario_norm → llave desde Swift ──
    consec_to_llave: Dict[str, str] = {}

    for _, r in df_swift_all.iterrows():
        llave = str(r.get("Llave", "") or "").strip()
        if not llave or llave.lower() in ("nan", "none", ""):
            continue
        for f in _extract_consecutivos_from_formulario(r.get("Formulario", "")):
            if f not in consec_to_llave:
                consec_to_llave[f] = llave

    if not consec_to_llave:
        LOGGER.info("PASO 5 → No hay formularios con llave en Swift_completos.")
        return

    LOGGER.info(
        f"PASO 5: {len(consec_to_llave)} formularios con llave. "
        f"Muestra: {list(consec_to_llave.items())[:5]}"
    )

    # ── Abrir workbook con openpyxl (preserva todos los formatos) ──
    from openpyxl import load_workbook

    try:
        wb = load_workbook(path)
    except PermissionError:
        raise PermissionError(
            f"\n{'='*60}\n"
            f"  No se pudo abrir origenDestino.xlsx — está abierto en Excel.\n"
            f"  Cerralo y volvé a ejecutar.\n"
            f"  Ruta: {path}\n"
            f"{'='*60}"
        )

    if config.SHEET_OD_ORIGEN not in wb.sheetnames:
        raise KeyError(
            f"No existe hoja '{config.SHEET_OD_ORIGEN}' en origenDestino.xlsx. "
            f"Hojas disponibles: {wb.sheetnames}"
        )

    ws = wb[config.SHEET_OD_ORIGEN]

    # ── Localizar columnas por header en fila 1 ─────────────────
    headers = {
        str(cell.value).replace("\u00A0", " ").strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }

    if config.OD2_COL_CONSECUTIVO not in headers:
        raise KeyError(
            f"No se encontró columna '{config.OD2_COL_CONSECUTIVO}' "
            f"en hoja '{config.SHEET_OD_ORIGEN}'. Headers: {list(headers.keys())}"
        )

    col_consec   = headers[config.OD2_COL_CONSECUTIVO]
    col_llave_od = headers.get(config.OD2_COL_LLAVE_OD)

    # Si la columna Llave Origen Destino no existe, crearla al final
    if col_llave_od is None:
        col_llave_od = ws.max_column + 1
        ws.cell(row=1, column=col_llave_od, value=config.OD2_COL_LLAVE_OD)
        LOGGER.info(f"Columna '{config.OD2_COL_LLAVE_OD}' creada en columna {col_llave_od}.")

    # ── Remover validación de datos de la columna Llave Origen Destino ──
    # La columna tiene un dropdown que referencia otra hoja. Si se deja activo,
    # Excel puede rechazar los valores escritos por openpyxl o marcar error al abrir.
    # Se elimina solo la validación de esa columna; el resto del sheet queda intacto.
    from openpyxl.utils import get_column_letter
    llave_col_letter = get_column_letter(col_llave_od)

    validaciones_a_conservar = []
    removidas = 0
    for dv in ws.data_validations.dataValidation:
        # Verificar si algún rango de esta validación incluye la columna llave
        toca_columna = any(
            llave_col_letter in str(rng)
            for rng in dv.sqref.ranges
        )
        if toca_columna:
            removidas += 1
        else:
            validaciones_a_conservar.append(dv)

    ws.data_validations.dataValidation = validaciones_a_conservar
    if removidas:
        LOGGER.info(
            f"PASO 5: {removidas} validación(es) de datos removida(s) "
            f"de columna '{config.OD2_COL_LLAVE_OD}' (columna {llave_col_letter})."
        )

    # ── Recorrer filas y escribir SOLO las celdas que cambian ───
    nuevos = actualizados = sin_match = 0

    for row_idx in range(2, ws.max_row + 1):
        raw_val = ws.cell(row=row_idx, column=col_consec).value

        if raw_val is None:
            continue

        # Normalizar consecutivo sin ceros iniciales (igual que formulario_str)
        if isinstance(raw_val, (int, float)):
            consec_norm = str(int(raw_val)).lstrip("0") or "0"
        else:
            sx = re.sub(r"\.0+$", "", str(raw_val).strip())
            m  = re.search(r"\d+", sx)
            consec_norm = (m.group(0).lstrip("0") or "0") if m else ""

        if not consec_norm:
            continue

        nueva_llave = consec_to_llave.get(consec_norm, "")
        if not nueva_llave:
            sin_match += 1
            continue

        cell_llave   = ws.cell(row=row_idx, column=col_llave_od)
        llave_actual = str(cell_llave.value or "").strip()

        if not llave_actual:
            cell_llave.value = nueva_llave
            nuevos += 1
        elif llave_actual != nueva_llave:
            cell_llave.value = nueva_llave
            actualizados += 1

    LOGGER.info(
        f"PASO 5 → nuevos={nuevos} | actualizados={actualizados} | sin_match={sin_match}"
    )

    # ── Guardar preservando todo el formato original ─────────────
    try:
        wb.save(path)
        LOGGER.info(
            f"origenDestino.xlsx guardado "
            f"(solo '{config.OD2_COL_LLAVE_OD}' modificado, resto intacto)."
        )
    except PermissionError:
        raise PermissionError(
            f"\n{'='*60}\n"
            f"  No se pudo guardar origenDestino.xlsx — está abierto en Excel.\n"
            f"  Cerralo y volvé a ejecutar.\n"
            f"  Ruta: {path}\n"
            f"{'='*60}"
        )


# =========================================================
# FUNCIÓN PRINCIPAL — llamada desde main.py
# =========================================================
def run_cruce_completo() -> Dict:
    """
    Ejecuta los 5 pasos del cruce.
    Retorna dict con estadísticas: formularios, llaves.
    """
    LOGGER.info("=== INICIO CRUCE FORMULARIOS + LLAVE ===")

    validate_input_files(
        config.XLSB_CUENTA_COM,
        config.SWIFT_COMPLETOS,
        config.ORIGEN_DESTINO,
        context="run_formulario",
    )

    stats = {"formularios": 0, "llaves": 0}

    # Paso 1 + 2: Leer y filtrar COM
    df_com           = read_com_sheet(config.XLSB_CUENTA_COM, config.SHEET_COM)
    df_com_filtrado  = filter_com_df(df_com)

    if df_com_filtrado.empty:
        LOGGER.warning("COM filtrada vacía. No se puede ejecutar cruce.")
        return stats

    # Preparar COM keys una sola vez
    com_keys = _build_com_keys(df_com_filtrado)

    # Leer Swift_completos
    df_v1 = read_sheet_safe(config.SWIFT_COMPLETOS, config.SHEET_V1, context="cruces")
    df_v2 = read_sheet_safe(config.SWIFT_COMPLETOS, config.SHEET_V2, context="cruces")

    # Forzar dtype object para columnas que recibirán strings
    for df in (df_v1, df_v2):
        for col in ("Formulario", "Llave"):
            if col in df.columns:
                df[col] = df[col].astype(object)

    # Paso 3: Formulario
    LOGGER.info("Paso 3: Cruce Formulario (COM → Swift por Nombre archivo + fecha)...")
    df_v1 = _update_formulario_for_sheet(df_v1, com_keys)
    df_v2 = _update_formulario_for_sheet(df_v2, com_keys)

    stats["formularios"] = int(
        df_v1["Formulario"].replace("", pd.NA).notna().sum()
        + df_v2["Formulario"].replace("", pd.NA).notna().sum()
    )

    # Paso 4: Llave
    LOGGER.info("Paso 4: Cruce Llave (origenDestino → Swift)...")
    od_map = _read_od_mapping(config.ORIGEN_DESTINO)
    df_v1  = _apply_llave_to_sheet(df_v1, od_map)
    df_v2  = _apply_llave_to_sheet(df_v2, od_map)

    stats["llaves"] = int(
        df_v1["Llave"].replace("", pd.NA).notna().sum()
        + df_v2["Llave"].replace("", pd.NA).notna().sum()
    )

    # Guardar Swift_completos
    write_sheets(
        config.SWIFT_COMPLETOS,
        {config.SHEET_V1: df_v1, config.SHEET_V2: df_v2},
        context="run_formulario",
    )

    # Paso 5: Llave Origen Destino
    LOGGER.info("Paso 5: Actualizando Llave Origen Destino en origenDestino.xlsx...")
    df_swift_all = pd.concat([df_v1, df_v2], ignore_index=True)
    _update_od_llave(config.ORIGEN_DESTINO, df_swift_all)

    LOGGER.info(
        f"=== FIN CRUCE ===  "
        f"Formularios={stats['formularios']} | Llaves={stats['llaves']}"
    )
    return stats


# =========================================================
# MAIN — ejecución standalone
# =========================================================
if __name__ == "__main__":
    run_cruce_completo()