"""
post_validacion_swift.py — Post validación y traslado de manuales corregidos

Pasos:
  1) Swift_manuales → Swift_completos
  2) Swift_completos → Acumulado_swift (APPEND sin duplicar por id)
  3) Swift_completos → Plantillas Bancolombia (REEMPLAZA datos desde fila 3,
     sin tocar fila 1 ni fila 2, preservando formatos con openpyxl directo)
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import config
from core.logger import get_logger
from core.excel_utils import read_sheet_safe, write_sheets, ensure_columns
from core.text_utils import build_nombre_personalizado, corregir_forma_societaria
from core.validators import validate_input_files

LOGGER = get_logger(__name__)

NOMBRE_PERSONALIZADO_LIMITE: int = 50


# =========================================================
# UTILIDADES INTERNAS
# =========================================================
def _recalcular_nombre_personalizado(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["Nombre personalizado"] = out.apply(
        lambda r: build_nombre_personalizado(r.get("Proveedor"), r.get("Receiver")),
        axis=1,
    )
    return out


def _es_completo(row) -> bool:
    def _has(v) -> bool:
        if v is None:
            return False
        if isinstance(v, float) and pd.isna(v):
            return False
        return str(v).strip().lower() not in ("", "nan", "none", "nat")
    return _has(row.get("Receiver")) and _has(row.get("Proveedor")) and _has(row.get("Amount"))


def _filtrar_completos(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["Estado"] = out.apply(lambda r: "Completo" if _es_completo(r) else "Incompleto", axis=1)
    return out.loc[out["Estado"] == "Completo"].copy()


def _ensure_id_column(df: pd.DataFrame, label: str) -> None:
    if not df.empty and "id" not in df.columns:
        raise KeyError(
            f"{label} no tiene columna 'id'. "
            "Esta columna es obligatoria para evitar duplicados."
        )


def _recortar_nombre_personalizado(valor, limite: int = NOMBRE_PERSONALIZADO_LIMITE):
    """
    Recorta Nombre personalizado al límite preservando el código SWIFT (última palabra).
    Elimina palabras del nombre de derecha a izquierda hasta que quepa en el límite.
    """
    if pd.isna(valor):
        return valor
    val = str(valor).strip()
    if len(val) <= limite:
        return val
    partes = val.rsplit(" ", 1)
    if len(partes) != 2:
        return val
    nombre, swift = partes[0], partes[1]
    palabras = nombre.split(" ")
    while len(" ".join(palabras) + " " + swift) > limite and len(palabras) > 1:
        palabras.pop()
    return " ".join(palabras) + " " + swift


# =========================================================
# HELPER DE RUTAS SEGÚN TIPO
# =========================================================
def _paths(tipo: str) -> dict:
    """
    Retorna las rutas correctas según el tipo de operación.
    tipo: "imp" | "exp"
    """
    if tipo == "exp":
        return {
            "manuales":  config.SWIFT_MANUALES_EXP,
            "completos": config.SWIFT_COMPLETOS_EXP,
            "acumulado": config.ACUMULADO_SWIFT_EXP,
            "plantilla": config.PLANTILLA_EXP,
        }
    return {
        "manuales":  config.SWIFT_MANUALES_IMP,
        "completos": config.SWIFT_COMPLETOS_IMP,
        "acumulado": config.ACUMULADO_SWIFT,
        "plantilla": config.PLANTILLA_IMP,
    }


# =========================================================
# FUNCIONES PÚBLICAS PARA main.py
# =========================================================
def contar_listos_en_manuales(tipo: str = "imp") -> tuple[int, int]:
    p = _paths(tipo)
    if not p["manuales"].exists():
        return 0, 0

    df_v1 = read_sheet_safe(p["manuales"], config.SHEET_V1, context="contar_manuales")
    df_v2 = read_sheet_safe(p["manuales"], config.SHEET_V2, context="contar_manuales")

    df_v1 = _recalcular_nombre_personalizado(df_v1) if not df_v1.empty else df_v1
    df_v2 = _recalcular_nombre_personalizado(df_v2) if not df_v2.empty else df_v2

    listos = sum([
        len(_filtrar_completos(df_v1)) if not df_v1.empty else 0,
        len(_filtrar_completos(df_v2)) if not df_v2.empty else 0,
    ])
    total = len(df_v1) + len(df_v2)
    return listos, total - listos


def ejecutar_mover_manuales(tipo: str = "imp") -> int:
    return _paso_1_mover_manuales(tipo)[2]


# =========================================================
# PASO 1) Swift_manuales → Swift_completos
# =========================================================
def _paso_1_mover_manuales(tipo: str = "imp") -> tuple[pd.DataFrame, pd.DataFrame, int]:
    p = _paths(tipo)
    validate_input_files(p["manuales"], context="post_validacion")

    man_v1 = read_sheet_safe(p["manuales"], config.SHEET_V1, context="manuales")
    man_v2 = read_sheet_safe(p["manuales"], config.SHEET_V2, context="manuales")

    man_v1 = _recalcular_nombre_personalizado(man_v1) if not man_v1.empty else man_v1
    man_v2 = _recalcular_nombre_personalizado(man_v2) if not man_v2.empty else man_v2

    man_v1_ok = _filtrar_completos(man_v1) if not man_v1.empty else pd.DataFrame()
    man_v2_ok = _filtrar_completos(man_v2) if not man_v2.empty else pd.DataFrame()

    comp_v1 = read_sheet_safe(p["completos"], config.SHEET_V1, context="completos")
    comp_v2 = read_sheet_safe(p["completos"], config.SHEET_V2, context="completos")

    for label, df in [
        ("Swift_manuales V1", man_v1_ok),
        ("Swift_manuales V2", man_v2_ok),
        ("Swift_completos V1", comp_v1),
        ("Swift_completos V2", comp_v2),
    ]:
        _ensure_id_column(df, label)

    if man_v1_ok.empty and man_v2_ok.empty:
        LOGGER.info("No hay registros completos nuevos en Swift_manuales.")
        return comp_v1, comp_v2, 0

    def _merge_sin_duplicar(comp: pd.DataFrame, nuevos: pd.DataFrame) -> tuple[pd.DataFrame, int]:
        if nuevos.empty:
            return comp, 0
        ids_existentes = set(comp["id"].astype(str)) if not comp.empty else set()
        to_add = nuevos.loc[~nuevos["id"].astype(str).isin(ids_existentes)].copy()
        if to_add.empty:
            return comp, 0
        to_add["Estado"] = "Completo"
        merged = pd.concat([comp, to_add], ignore_index=True) if not comp.empty else to_add
        return merged, len(to_add)

    new_comp_v1, added_v1 = _merge_sin_duplicar(comp_v1, man_v1_ok)
    new_comp_v2, added_v2 = _merge_sin_duplicar(comp_v2, man_v2_ok)
    total_movidos = added_v1 + added_v2

    if total_movidos > 0:
        write_sheets(
            p["completos"],
            {config.SHEET_V1: new_comp_v1, config.SHEET_V2: new_comp_v2},
            context=f"paso_1_manuales_{tipo}",
        )
        LOGGER.info(
            f"PASO 1 OK [{tipo.upper()}] → movidos a Swift_completos: "
            f"V1={added_v1} | V2={added_v2} | total={total_movidos}"
        )
    else:
        LOGGER.info("PASO 1: Todos los registros completos ya existían en Swift_completos.")

    return new_comp_v1, new_comp_v2, total_movidos


# =========================================================
# PASO 2) Swift_completos → Acumulado_swift (APPEND)
# =========================================================
def _paso_2_acumulado(df_v1: pd.DataFrame, df_v2: pd.DataFrame, tipo: str = "imp") -> None:
    p = _paths(tipo)
    df_v1 = df_v1.copy()
    df_v2 = df_v2.copy()
    df_v1["Versión"] = "V1"
    df_v2["Versión"] = "V2"

    df_all = pd.concat([df_v1, df_v2], ignore_index=True)

    if "id" not in df_all.columns:
        raise KeyError("Swift_completos no tiene columna 'id'. Es obligatoria para el acumulado.")

    ahora = datetime.now().strftime("%Y-%m-%d %H:%M")
    df_all["Fecha Control"] = ahora

    df_all = ensure_columns(df_all, config.ACUMULADO_COLS)
    df_new = df_all[config.ACUMULADO_COLS].copy()

    if p["acumulado"].exists():
        df_old = read_sheet_safe(
            p["acumulado"], config.SHEET_ACUMULADO, context="acumulado"
        )
        df_old = ensure_columns(df_old, config.ACUMULADO_COLS)

        ids_old = set(df_old["id"].astype(str))
        df_to_add = df_new.loc[~df_new["id"].astype(str).isin(ids_old)].copy()

        if df_to_add.empty:
            LOGGER.info("PASO 2: Sin registros nuevos para acumulado (todos ya existen).")
            return

        df_final = pd.concat([df_old, df_to_add], ignore_index=True)
        LOGGER.info(
            f"PASO 2 OK [{tipo.upper()}] → Acumulado: agregados={len(df_to_add)} | total={len(df_final)}"
        )
    else:
        df_final = df_new.copy()
        LOGGER.info(f"PASO 2: Acumulado {tipo.upper()} creado con {len(df_final)} registros.")

    write_sheets(
        p["acumulado"],
        {config.SHEET_ACUMULADO: df_final},
        context=f"acumulado_{tipo}",
    )


# =========================================================
# PASO 3) Swift_completos → Plantillas Bancolombia (REEMPLAZA)
# =========================================================
def _preparar_df_para_plantilla(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    if "Estado" in out.columns:
        out = out.loc[out["Estado"] == "Completo"].copy()

    # Corregir formas societarias en Proveedor y Nombre personalizado
    # (función importada desde core.text_utils — misma implementación que en run_formulario)
    for col in ("Proveedor", "Nombre personalizado"):
        if col in out.columns:
            out[col] = out[col].apply(
                lambda v: corregir_forma_societaria(str(v)) if pd.notna(v) and str(v).strip() else v
            )

    # Deduplicar por "Nombre personalizado" — Bancolombia rechaza duplicados
    antes = len(out)
    if "Nombre personalizado" in out.columns:
        out = out.drop_duplicates(subset=["Nombre personalizado"]).reset_index(drop=True)
    else:
        out = out.drop_duplicates().reset_index(drop=True)
    duplicados = antes - len(out)
    if duplicados:
        LOGGER.info(f"Plantilla: {duplicados} registros duplicados eliminados.")

    if "Nombre personalizado" in out.columns:
        originales = out["Nombre personalizado"].copy()
        out["Nombre personalizado"] = out["Nombre personalizado"].apply(
            _recortar_nombre_personalizado
        )
        recortados = (out["Nombre personalizado"] != originales).sum()
        if recortados:
            LOGGER.info(
                f"Plantilla: {recortados} nombres recortados a {NOMBRE_PERSONALIZADO_LIMITE} chars."
            )

    return out


def _crear_plantilla_desde_base(template_path: Path) -> None:
    """
    Crea una plantilla Bancolombia vacía (fila 1 + fila 2) cuando no existe.

    Estrategia en orden de prioridad:
      1. Copiar la plantilla IMP equivalente (V1 o V2) y limpiar datos
      2. Si tampoco existe la IMP, construir la estructura mínima desde cero
    """
    import shutil
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    # Asegurar que la carpeta plantillas existe
    template_path.parent.mkdir(parents=True, exist_ok=True)

    # Usar la plantilla IMP como base para cualquier tipo
    base_imp = config.PLANTILLA_IMP

    if base_imp.exists():
        # ── Opción 1: copiar IMP y limpiar datos ──
        shutil.copy2(base_imp, template_path)
        wb = load_workbook(template_path)
        ws = wb.active
        if ws.max_row >= 3:
            ws.delete_rows(3, ws.max_row - 2)
        wb.save(template_path)
        LOGGER.info(
            f"Plantilla autogenerada desde base IMP: {template_path.name}"
        )
        return

    # ── Opción 2: construir desde cero ──
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos Origen y Destino"

    # Fila 1: título combinado A1:F1
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Datos Origen y Destino"
    title_cell.font = Font(name="Calibri", size=20, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 41.25

    # Fila 2: encabezados
    headers = [
        "Cuenta compensación",
        "Nombre / Razón social",
        "País",
        "Ciudad",
        "Nombre personalizado",
        "Observaciones",
    ]
    col_widths = [52.43, 46.0, 27.43, 20.57, 35.71, 148.14]
    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(name="Calibri", bold=True)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    # Hoja Parametros con lista de países (requerida por Bancolombia)
    ws_p = wb.create_sheet("Parametros")
    ws_p.cell(row=1, column=1, value="Paises")
    paises = [
        "AFGANISTAN","ALBANIA","ALEMANIA","ANDORRA","ANGOLA","ANGUILA",
        "ANTIGUA Y BARBUDA","ARABIA SAUDITA","ARGELIA","ARGENTINA","ARMENIA",
        "ARUBA","AUSTRALIA","AUSTRIA","AZERBAIYAN","BAHAMAS","BAHREIN",
        "BANGLADESH","BARBADOS","BELARUS","BELGICA","BELICE","BENIN",
        "BERMUDAS","BUTAN","BOLIVIA","BOSNIA HERZEGOVINA","BOTSWANA","BRASIL",
        "BRUNEI","BULGARIA","BURKINA FASO","BURUNDI","CABO VERDE","CAMBOYA",
        "CAMERUN","CANADA","CHAD","CHILE","CHINA","CHIPRE","COLOMBIA",
        "COMORAS","CONGO","COREA DEL NORTE","COREA DEL SUR","COSTA DE MARFIL",
        "COSTA RICA","CROACIA","CUBA","CURACAO","DINAMARCA","DJIBOUTI",
        "DOMINICA","ECUADOR","EGIPTO","EL SALVADOR","EMIRATOS ARABES UNIDOS",
        "ERITREA","ESLOVAQUIA","ESLOVENIA","ESPAÑA","ESTADOS UNIDOS",
        "ESTONIA","ETIOPIA","FILIPINAS","FINLANDIA","FRANCIA","GABON",
        "GAMBIA","GEORGIA","GHANA","GIBRALTAR","GRECIA","GRENADA","GUATEMALA",
        "GUINEA","GUINEA ECUATORIAL","GUINEA-BISAU","GUYANA","HAITI",
        "HONDURAS","HONG KONG","HUNGRIA","INDIA","INDONESIA","IRAN","IRAQ",
        "IRLANDA","ISLANDIA","ISLAS CAIMAN","ISLAS COOK","ISLAS FEROE",
        "ISLAS MARSHALL","ISLAS SALOMON","ISLAS TURCAS Y CAICOS","ISLAS VIRGENES",
        "ISRAEL","ITALIA","JAMAICA","JAPON","JORDANIA","KAZAJSTAN","KENIA",
        "KIRGUISTAN","KIRIBATI","KUWAIT","LAOS","LESOTO","LETONIA","LIBANO",
        "LIBERIA","LIBIA","LIECHTENSTEIN","LITUANIA","LUXEMBURGO","MACAO",
        "MADAGASCAR","MALASIA","MALAWI","MALDIVAS","MALI","MALTA",
        "MARRUECOS","MAURICIO","MAURITANIA","MEXICO","MOLDAVIA","MONACO",
        "MONGOLIA","MONTENEGRO","MONTSERRAT","MOZAMBIQUE","MYANMAR","NAMIBIA",
        "NAURU","NEPAL","NICARAGUA","NIGER","NIGERIA","NORUEGA",
        "NUEVA ZELANDA","OMAN","PAISES BAJOS","PAKISTAN","PALAU","PANAMA",
        "PAPUA NUEVA GUINEA","PARAGUAY","PERU","POLONIA","PORTUGAL",
        "PUERTO RICO","QATAR","REINO UNIDO","REPUBLICA CENTROAFRICANA",
        "REPUBLICA CHECA","REPUBLICA DEMOCRATICA DEL CONGO",
        "REPUBLICA DOMINICANA","RUMANIA","RUSIA","RWANDA","SAMOA",
        "SAN CRISTOBAL Y NIEVES","SAN MARINO","SAN VICENTE Y LAS GRANADINAS",
        "SANTA LUCIA","SENEGAL","SERBIA","SEYCHELLES","SIERRA LEONA",
        "SINGAPUR","SIRIA","SOMALIA","SRI LANKA","SUDAFRICA","SUDAN",
        "SURINAM","SUECIA","SUIZA","TAILANDIA","TAIWAN","TANZANIA",
        "TIMOR ORIENTAL","TOGO","TONGA","TRINIDAD Y TOBAGO","TUNEZ",
        "TURKMENISTAN","TURQUIA","TUVALU","UCRANIA","UGANDA","URUGUAY",
        "UZBEKISTAN","VANUATU","VENEZUELA","VIETNAM","YEMEN","YIBUTI",
        "ZAMBIA","ZIMBABUE",
    ]
    for i, pais in enumerate(paises, start=2):
        ws_p.cell(row=i, column=1, value=pais)

    wb.save(template_path)
    LOGGER.info(f"Plantilla creada desde cero: {template_path.name}")


def _write_bancolombia_template(df_source: pd.DataFrame, template_path: Path) -> None:
    """
    Escribe datos en la plantilla Bancolombia usando openpyxl directamente.

    - Fila 1: logo/título → intocable
    - Fila 2: encabezados → intocable
    - Fila 3+: datos → REEMPLAZA completamente preservando formato de celdas vacías

    Al usar load_workbook + wb.save() en lugar de pd.ExcelWriter,
    se preservan todos los formatos, bordes y estilos del template.
    """
    if not template_path.exists():
        _crear_plantilla_desde_base(template_path)

    df_write = _preparar_df_para_plantilla(df_source)

    # ── Abrir con openpyxl (preserva fila 1 y fila 2 intactas) ──
    wb = load_workbook(template_path)
    ws = wb.active

    # Mapear encabezados de fila 2 → {nombre: col_number}
    header_map = {
        str(cell.value).replace("\u00A0", " ").strip(): cell.column
        for cell in ws[2]
        if cell.value is not None
    }

    required_headers = [
        "Cuenta compensación",
        "Nombre / Razón social",
        "País",
        "Ciudad",
        "Nombre personalizado",
    ]
    missing = [h for h in required_headers if h not in header_map]
    if missing:
        raise KeyError(
            f"Plantilla {template_path.name} no tiene encabezados requeridos "
            f"en fila 2: {missing}"
        )

    # Limpiar datos desde fila 3 (sin tocar filas 1 y 2)
    start_row = 3
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)

    if df_write.empty:
        wb.save(template_path)
        LOGGER.info(f"PASO 3: Plantilla limpia (sin registros): {template_path.name}")
        return

    # Escribir celda a celda (solo las 5 columnas requeridas)
    col_cc  = header_map["Cuenta compensación"]
    col_rs  = header_map["Nombre / Razón social"]
    col_pa  = header_map["País"]
    col_ci  = header_map["Ciudad"]
    col_np  = header_map["Nombre personalizado"]

    for current_row, (_, r) in enumerate(df_write.iterrows(), start=start_row):
        ws.cell(row=current_row, column=col_cc).value = config.CUENTA_COMPENSACION
        ws.cell(row=current_row, column=col_rs).value = r.get("Proveedor", "")
        ws.cell(row=current_row, column=col_pa).value = r.get("Pais", "")
        ws.cell(row=current_row, column=col_ci).value = r.get("Ciudad", "")
        ws.cell(row=current_row, column=col_np).value = r.get("Nombre personalizado", "")

    try:
        wb.save(template_path)
    except PermissionError:
        raise PermissionError(
            f"\n{'='*60}\n"
            f"  No se pudo guardar {template_path.name} — está abierto en Excel.\n"
            f"  Cerralo y volvé a ejecutar.\n"
            f"  Ruta: {template_path}\n"
            f"{'='*60}"
        )

    LOGGER.info(
        f"PASO 3 OK → Plantilla actualizada: {template_path.name} "
        f"({len(df_write)} registros desde fila 3)"
    )


def _paso_3_plantillas(df_v1: pd.DataFrame, df_v2: pd.DataFrame, tipo: str = "imp") -> int:
    """
    Concatena V1 + V2 y genera una sola plantilla Bancolombia por tipo.
    Retorna el total de registros escritos.
    """
    p = _paths(tipo)
    # Unir ambas versiones — el origen del PDF (V1 o V2) es irrelevante
    # para Bancolombia, solo importa el contenido
    df_todos = pd.concat([df_v1, df_v2], ignore_index=True)
    _write_bancolombia_template(df_todos, p["plantilla"])
    return len(_preparar_df_para_plantilla(df_todos))


# =========================================================
# FUNCIÓN PRINCIPAL — llamada desde main.py
# =========================================================
def run_post_validacion(tipo: str = "imp") -> dict:
    """
    Paso 1: manuales → completos
    Paso 2: completos → acumulado
    NO genera plantilla (eso es run_generar_plantilla).
    """
    LOGGER.info(f"=== INICIO POST VALIDACIÓN [{tipo.upper()}] ===")

    df_comp_v1, df_comp_v2, total_movidos = _paso_1_mover_manuales(tipo)

    if not df_comp_v1.empty or not df_comp_v2.empty:
        _paso_2_acumulado(df_comp_v1, df_comp_v2, tipo)
    else:
        LOGGER.info("PASO 2: Swift_completos vacío, se omite acumulado.")

    LOGGER.info(f"=== FIN POST VALIDACIÓN [{tipo.upper()}] ===  Movidos={total_movidos}")
    return {"movidos": total_movidos}


def run_generar_plantilla(tipo: str = "imp") -> dict:
    """
    Paso 3: Lee Swift_completos (V1 + V2) y genera UNA SOLA plantilla
    Bancolombia por tipo:
      - IMP → Plantilla_imp.xlsx
      - EXP → Plantilla_exp.xlsx
    Independiente del paso de manuales — se puede ejecutar en cualquier momento.
    """
    LOGGER.info(f"=== INICIO GENERAR PLANTILLA [{tipo.upper()}] ===")

    p = _paths(tipo)

    if not p["completos"].exists():
        raise FileNotFoundError(
            f"No existe Swift_completos. Ejecuta primero la extracción OCR."
        )

    df_v1 = read_sheet_safe(p["completos"], config.SHEET_V1, context="plantilla")
    df_v2 = read_sheet_safe(p["completos"], config.SHEET_V2, context="plantilla")

    total = _paso_3_plantillas(df_v1, df_v2, tipo)

    LOGGER.info(
        f"=== FIN GENERAR PLANTILLA [{tipo.upper()}] === "
        f"Archivo: {p['plantilla'].name} | Registros={total}"
    )
    return {"registros": total, "archivo": str(p["plantilla"])}